from django.shortcuts import render, redirect
from django.http import HttpResponse
from ..models import *
import openpyxl
from openpyxl.utils import get_column_letter
import os
from django.conf import settings
import subprocess
from urllib.parse import urlparse
import pandas as pd
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.http import JsonResponse
from django.db.models import Prefetch

# Vista para descargar la base de datos
def descargar_base_datos(request):
    # Crear un nombre de archivo basado en la fecha actual
    fecha_actual = timezone.now().strftime('%Y-%m-%d_%H-%M-%S')

    # Comprobamos el tipo de base de datos
    db_engine = settings.DATABASES['default']['ENGINE']
    
    if 'sqlite3' in db_engine:
        # Para SQLite, usamos el archivo .sqlite3 directamente
        db_path = settings.DATABASES['default']['NAME']
        filename = f"backup_{fecha_actual}.sqlite3"
        
        # Abrimos el archivo y lo enviamos como respuesta
        with open(db_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type="application/x-sqlite3")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    
    elif 'postgresql' in db_engine:
        # Obtener la URL de conexión completa de la base de datos
        db_url = "postgresql://bomberos_dbs_user:sB0cG00nvzS11aSYoROt2wNSWPnSIDyR@dpg-ctct69btq21c7380aogg-a.oregon-postgres.render.com/bomberos_dbs"
        
        url_parsed = urlparse(db_url)

        # Extraer información de la URL
        db_name = url_parsed.path[1:]  # Remueve el primer '/' de la ruta para obtener solo el nombre
        db_user = url_parsed.username
        db_password = url_parsed.password
        db_host = url_parsed.hostname

        # Configurar el nombre del archivo de respaldo
        filename = f"backup_{fecha_actual}.sql"

        # Configurar el comando pg_dump
       # Cambiar el comando pg_dump para exportar en formato de texto plano
        dump_cmd = [
            "pg_dump",
            "-h", db_host,
            "-U", db_user,
            "-d", db_name,
            "-F", "p"  # Cambiado a texto plano
        ]


        # Establecer la contraseña en la variable de entorno y ejecutar el comando
        os.environ['PGPASSWORD'] = db_password
        with subprocess.Popen(dump_cmd, stdout=subprocess.PIPE) as proc:
            response = HttpResponse(proc.stdout, content_type="application/sql")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response    
    else:
        # Otros motores de base de datos
        return HttpResponse("Motor de base de datos no compatible", status=400)

# Api para crear el excel de exportacion de la tabla
def generar_excel_personal(request):
    # Crear un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Personal"

    # Agregar encabezados a la primera fila
    encabezados = [
        "Nombres", "Apellidos", "Jerarquia", "Cargo", 
        "Cedula", "Sexo", "Contrato", "Estado"
    ]
    hoja.append(encabezados)

    # Obtener datos de los procedimientos
    procedimientos = Personal.objects.exclude(id__in=[0, 4])

    # Agregar datos a la hoja
    for procedimiento in procedimientos:
        # Agregar la fila de datos
        hoja.append([
            procedimiento.nombres,
            procedimiento.apellidos,
            procedimiento.jerarquia,
            procedimiento.cargo,
            procedimiento.cedula,
            procedimiento.sexo,
            procedimiento.rol,
            procedimiento.status,
        ])

    # Ajustar el ancho de las columnas
    for column in hoja.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
        hoja.column_dimensions[get_column_letter(column[0].column)].width = max_length

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=personal.xlsx"
    workbook.save(response)
    return response

def generar_excel_operaciones(request):
    division = 2

    mes_excel = request.GET.get('mes') 

    # Extraer año y mes
    año, mes_num = mes_excel.split("-")

    # Obtener datos de los procedimientos con select_related y prefetch_related
    procedimientos = Procedimientos.objects.filter(id_division=division, fecha__year=año, fecha__month=mes_num).select_related(
        'id_solicitante', 'id_jefe_comision', 'id_municipio', 'id_parroquia', 'id_tipo_procedimiento'
    ).prefetch_related(
        Prefetch("rescate_set", 
             queryset=Rescate.objects.prefetch_related(
                 Prefetch("rescate_persona_set", to_attr="personas_data"),
                 Prefetch("rescate_animal_set", to_attr="animales_data")
             ), 
             to_attr="rescate_data"),
        Prefetch("abastecimiento_agua_set", to_attr="agua_data"), # Ya
        Prefetch("apoyo_unidades_set", to_attr="apoyo_data"), # Ya
        Prefetch("guardia_prevencion_set", to_attr="guard_data"), # Ys
        Prefetch("falsa_alarma_set", to_attr="falsa_data"), # YA
        Prefetch("atenciones_paramedicas_set", 
             queryset=Atenciones_Paramedicas.objects.prefetch_related(
                 Prefetch("emergencias_medicas_set", 
                          queryset=Emergencias_Medicas.objects.prefetch_related(
                              Prefetch("traslado_set", to_attr="traslados_data")
                          ), 
                          to_attr="emergencias_data"),
                 Prefetch("accidentes_transito_set", 
                          queryset=Accidentes_Transito.objects.prefetch_related(
                              Prefetch("detalles_vehiculos_accidente_set", to_attr="vehiculos_data"),
                              Prefetch("lesionados_set", 
                                       queryset=Lesionados.objects.prefetch_related(
                                           Prefetch("traslado_accidente_set", to_attr="traslados_accidente_data")
                                       ), 
                                       to_attr="lesionados_data")
                          ), 
                          to_attr="accidentes_data")
             ), 
             to_attr="atenciones_data"), # Ya
        Prefetch("servicios_especiales_set", to_attr="especial_data"), # Ya
        Prefetch("incendios_set", 
             queryset=Incendios.objects.prefetch_related(
                 Prefetch("retencion_preventiva_incendios_set", to_attr="retencion_data"),
                 Prefetch("persona_presente_set", to_attr="personas_incendio_data"),
                 Prefetch("detalles_vehiculos_set", to_attr="vehiculos_incendio_data")
             ), 
             to_attr="incendios_data"), # Ya
        Prefetch("mitigacion_riesgos_set", to_attr="mitigacion_data"), # Ya
        Prefetch("evaluacion_riesgo_set", 
             queryset=Evaluacion_Riesgo.objects.prefetch_related(
                 Prefetch("persona_presente_eval_set", to_attr="personas_eval_data")
             ), 
             to_attr="evaluacion_data"),# Ya
        Prefetch("puesto_avanzada_set", to_attr="puesto_data"), # Ya
        Prefetch("artificios_pirotecnicos_set",
        queryset=Artificios_Pirotecnicos.objects.prefetch_related(
            Prefetch(
                "incendios_art_set",
                queryset=Incendios_Art.objects.prefetch_related(
                    Prefetch("persona_presente_art_set", to_attr="personas_presentes_data"),
                    Prefetch("detalles_vehiculos_art_set", to_attr="vehiculos_data")
                ),
                to_attr="incendios_data"
            ),
            Prefetch("lesionados_art_set", to_attr="lesionados_data"),
            Prefetch("fallecidos_art_set", to_attr="fallecidos_data"),
        ),
        to_attr="artificios_data"), # Ya
        Prefetch("comisiones_set", 
             queryset=Comisiones.objects.select_related("comision"), 
             to_attr="comisiones_data"), # Ya
        Prefetch("atendido_no_efectuado_set", to_attr="atendido_data"), # Ya
        Prefetch("despliegue_seguridad_set", to_attr="despliegue_data"), # Ya
        Prefetch("fallecidos_set", to_attr="fallecido_data"), # Ya
    )

    datos = []

    # Agregar datos a la lista
    for procedimiento in procedimientos:
        # Obtener solicitante y jefe de comisión

        if procedimiento.id_solicitante.apellidos == "Externo":
            solicitante = procedimiento.solicitante_externo
        else:
            solicitante = f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}"


        jefe_comision = (f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}")

        # Detalles y personas
        personas_presentes = []
        detalles = []
        descripcion = []
        material_utilizado = []
        status = []
        traslados = []
        vehiculos = []
        comisiones_presentes = []
        retencion_preventiva = []

        # Agregar datos de comisiones presentes
        for comision in procedimiento.comisiones_data:
            comisiones_presentes.append(f"({comision.comision.tipo_comision} - {comision.nombre_oficial} {comision.apellido_oficial} ({comision.cedula_oficial}) - Unidad {comision.nro_unidad} - Cuadrante {comision.nro_cuadrante})")

        # Otros conjuntos relacionados (agregar más si aplica)
        for agua in procedimiento.agua_data:
            detalles.append(f"{agua.id_tipo_servicio.nombre_institucion} {agua.ltrs_agua} litros - {agua.personas_atendidas} Personas Atendidas")
            personas_presentes.append(f"{agua.nombres} {agua.apellidos} ({agua.cedula})")
            descripcion.append(agua.descripcion)
            material_utilizado.append(agua.material_utilizado)
            status.append(agua.status)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for apoyo in procedimiento.apoyo_data:
            detalles.append(f"{apoyo.id_tipo_apoyo.tipo_apoyo}: {apoyo.unidad_apoyada}")
            descripcion.append(apoyo.descripcion)
            material_utilizado.append(apoyo.material_utilizado)
            status.append(apoyo.status)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for guardia in procedimiento.guard_data:
            detalles.append(guardia.id_motivo_prevencion.motivo)
            descripcion.append(guardia.descripcion)
            material_utilizado.append(guardia.material_utilizado)
            status.append(guardia.status)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for atendido in procedimiento.atendido_data:
            descripcion.append(atendido.descripcion)
            material_utilizado.append(atendido.material_utilizado)
            status.append(atendido.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for despliegue in procedimiento.despliegue_data:
            detalles.append(despliegue.motivo_despliegue.motivo)
            descripcion.append(despliegue.descripcion)
            material_utilizado.append(despliegue.material_utilizado)
            status.append(despliegue.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for alarma in procedimiento.falsa_data:
            detalles.append(alarma.motivo_alarma.motivo)
            descripcion.append(alarma.descripcion)
            material_utilizado.append(alarma.material_utilizado)
            status.append(alarma.status)
        
        for atencion in procedimiento.atenciones_data:
            # Agregar detalles del tipo de atención paramédica

            # Agregar emergencias médicas si existen
            for emergencia in atencion.emergencias_data:
                detalles.append(atencion.tipo_atencion)
                personas_presentes.append(f" {emergencia.nombres} {emergencia.apellidos} ({emergencia.cedula}) {emergencia.edad} años - {emergencia.sexo} [{emergencia.idx}]")
                descripcion.append(emergencia.descripcion)
                material_utilizado.append(emergencia.material_utilizado)
                status.append(emergencia.status)

                # Agregar traslados si existen
                for traslado in emergencia.traslados_data:
                    traslados.append(f"Traslado: {traslado.hospital_trasladado} - {traslado.medico_receptor} - {traslado.mpps_cmt}")

              # Agregar datos de accidentes de tránsito
            for accidente in atencion.accidentes_data:
                detalles.append(f"{atencion.tipo_atencion}: {accidente.tipo_de_accidente.tipo_accidente} - {accidente.cantidad_lesionados} Lesionados")
                material_utilizado.append(accidente.material_utilizado)
                status.append(accidente.status)

                # Agregar detalles de vehículos involucrados
                for vehiculo in accidente.vehiculos_data:
                    vehiculos.append(f"({vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas})")

                # Agregar datos de lesionados
                for lesionado in accidente.lesionados_data:
                    personas_presentes.append(f"({lesionado.nombres} {lesionado.apellidos} {lesionado.cedula} - {lesionado.edad} años{lesionado.sexo} [{lesionado.idx}])")
                    descripcion.append(f"({lesionado.descripcion})")

                    # Agregar traslados asociados a lesionados
                    for traslado_acc in lesionado.traslados_accidente_data:
                        traslados.append(f"Traslado: {traslado_acc.hospital_trasladado} - {traslado_acc.medico_receptor} - {traslado_acc.mpps_cmt}")
 
        # Otros conjuntos relacionados (agregar más si aplica)
        for especial in procedimiento.especial_data:
            detalles.append(especial.tipo_servicio.serv_especiales)
            descripcion.append(especial.descripcion)
            material_utilizado.append(especial.material_utilizado)
            status.append(especial.status)

        for rescate in procedimiento.rescate_data:
            material_utilizado.append(rescate.material_utilizado)
            status.append(rescate.status)

            # Agregar personas presentes si existen
            for persona in rescate.personas_data:
                # Agregar detalles del tipo de rescate
                detalles.append(rescate.tipo_rescate.tipo_rescate)
                material_utilizado.append(rescate.material_utilizado)
                status.append(rescate.status)
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula} - {persona.edad} años - {persona.sexo})")
                descripcion.append(persona.descripcion)

            # Agregar animales rescatados si existen
            for animal in rescate.animales_data:
                detalles.append(f"{rescate.tipo_rescate.tipo_rescate}: {animal.especie}")
                descripcion.append(animal.descripcion)
     
        # Otros conjuntos relacionados (agregar más si aplica)
        for incendio in procedimiento.incendios_data:
            # Detalles del incendio
            detalles.append(incendio.id_tipo_incendio.tipo_incendio)
            descripcion.append(incendio.descripcion)
            material_utilizado.append(incendio.material_utilizado)
            status.append(incendio.status)

            # Agregar Retenciones Preventivas (GLP)
            for retencion in incendio.retencion_data:

                retencion_preventiva.append(f"{retencion.tipo_cilindro} - {retencion.capacidad} - {retencion.serial} - {retencion.nro_constancia_retencion} - Propietario: {retencion.nombre} {retencion.apellidos} ({retencion.cedula})")

            # Agregar Personas Presentes en el Incendio
            for persona in incendio.personas_incendio_data:
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula} - {persona.edad} años)")

            # Agregar Detalles de Vehículos Relacionados
            for vehiculo in incendio.vehiculos_incendio_data:
                vehiculos.append(f"{vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas}")

        # Agregar Fallecidos
        for fallecido in procedimiento.fallecido_data:
            detalles.append(fallecido.motivo_fallecimiento)
            personas_presentes.append(f"{fallecido.nombres} {fallecido.apellidos} ({fallecido.cedula}) - {fallecido.edad} años - {fallecido.sexo}") 
            descripcion.append(fallecido.descripcion)
            material_utilizado.append(fallecido.material_utilizado)
            status.append(fallecido.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for mitigacion in procedimiento.mitigacion_data:
            detalles.append(mitigacion.id_tipo_servicio.tipo_servicio)
            descripcion.append(mitigacion.descripcion)
            material_utilizado.append(mitigacion.material_utilizado)
            status.append(mitigacion.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for evaluacion in procedimiento.evaluacion_data:
            detalles.append(f"{evaluacion.id_tipo_riesgo.tipo_riesgo}: {evaluacion.tipo_estructura}")
            descripcion.append(evaluacion.descripcion)
            material_utilizado.append(evaluacion.material_utilizado)
            status.append(evaluacion.status)

            # Personas presentes en la evaluación
            for persona in evaluacion.personas_eval_data:
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula}) - {persona.telefono}")

        # Otros conjuntos relacionados (agregar más si aplica)
        for artificio in procedimiento.artificios_data:
            # Información básica del artificio
            detalles.append(f"{artificio.tipo_procedimiento.tipo} - {artificio.nombre_comercio} - {artificio.rif_comerciante}")

            for incendio in artificio.incendios_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}: {incendio.id_tipo_incendio.tipo_incendio}")
                descripcion.append(incendio.descripcion)
                material_utilizado.append(incendio.material_utilizado)
                status.append(incendio.status)

                for persona in incendio.personas_presentes_data:
                    personas_presentes.append(f"{persona.nombres} {persona.apellidos} ({persona.cedula}) - {persona.edad} años")
                    
                for vehiculo in incendio.vehiculos_data:
                    vehiculos.append(f"{vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas}")


            for lesionado in artificio.lesionados_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}")
                personas_presentes.append(f"{lesionado.nombres} {lesionado.apellidos} ({lesionado.cedula}) - {lesionado.edad} años - {lesionado.sexo}")
                descripcion.append(lesionado.descripcion)
                status.append(lesionado.status)

            for fallecido in artificio.fallecidos_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}: {fallecido.motivo_fallecimiento}")
                personas_presentes.append(f"{fallecido.nombres} {fallecido.apellidos} ({fallecido.cedula}) - {fallecido.edad} años - {fallecido.sexo}")
                descripcion.append(fallecido.descripcion)
                status.append(fallecido.status)
                material_utilizado.append(fallecido.material_utilizado)


        # Combinar datos
        detalles_str = " -- ".join(detalles) or " "
        personas_str = " -- ".join(personas_presentes) or " "
        descripcion_str = " -- ".join(descripcion) or " "
        material_utilizado_str = " -- ".join(material_utilizado) or " "
        status_str = " -- ".join(status) or " "
        traslados_str = " -- ".join(traslados) or " "
        vehiculos_str = " -- ".join(vehiculos) or " "
        comisiones_presentes_str = " -- ".join(comisiones_presentes) or " "
        retencion_preventiva_str = " -- ".join(retencion_preventiva) or " "

        # Agregar datos al JSON
        datos.append({
            "division": procedimiento.id_division.division or " ",
            "solicitante": solicitante,
            "jefe_comision": jefe_comision,
            "municipio": procedimiento.id_municipio.municipio or " ",
            "parroquia": procedimiento.id_parroquia.parroquia or " ",
            "fecha": procedimiento.fecha or " ",
            "hora": procedimiento.hora or " ",
            "direccion": procedimiento.direccion or " ",
            "tipo_procedimiento": procedimiento.id_tipo_procedimiento.tipo_procedimiento or " ",
            "detalles": detalles_str,
            "personas_presentes": personas_str,
            "descripcion": descripcion_str,
            "material_utilizado": material_utilizado_str,
            "status": status_str,
            "traslados": traslados_str,
            "vehiculos": vehiculos_str,
            "comisiones": comisiones_presentes_str,
            "retencion_preventiva": retencion_preventiva_str,
        })

    return JsonResponse(datos, safe=False)

def generar_excel_enfermeria(request):
    division = 6

    mes_excel = request.GET.get('mes') 

    # Extraer año y mes
    año, mes_num = mes_excel.split("-")

    # Optimización de consultas con select_related y prefetch_related
    procedimientos = (
        Procedimientos.objects.filter(id_division=division, fecha__year=año, fecha__month=mes_num)
        .select_related(
            "id_division",
            "id_municipio",
            "id_parroquia",
            "id_tipo_procedimiento",
        )
        .prefetch_related(
            Prefetch(
                "detalles_enfermeria_set",
                to_attr="enfermeria_data"
            )
        )
    )

    datos = []

    for procedimiento in procedimientos:
        # Inicializar listas para recolectar información
        personas_presentes = []
        descripcion = []
        material_utilizado = []
        status = []
        traslados = []

        # Procesar los detalles de enfermería
        for enfer in procedimiento.enfermeria_data:
            personas_presentes.append(f"{enfer.nombre} {enfer.apellido} ({enfer.cedula}) {enfer.edad} años {enfer.sexo} - [{enfer.telefono}]")
            descripcion.append(enfer.descripcion)
            material_utilizado.append(enfer.material_utilizado)
            status.append(enfer.status)

        # Combinar datos en cadenas
        personas_str = " -- ".join(personas_presentes) or "N/A"
        descripcion_str = " -- ".join(descripcion) or "N/A"
        material_utilizado_str = " -- ".join(material_utilizado) or "N/A"
        status_str = " -- ".join(status) or "N/A"
        traslados_str = " -- ".join(traslados) or "N/A"  # No se llenan traslados, conservar como "N/A"

        # Agregar datos procesados al resultado final
        datos.append(
            {
                "division": procedimiento.id_division.division,
                "dependencia": procedimiento.dependencia,
                "encargado": procedimiento.solicitante_externo,
                "municipio": procedimiento.id_municipio.municipio,
                "parroquia": procedimiento.id_parroquia.parroquia,
                "fecha": procedimiento.fecha,
                "hora": procedimiento.hora,
                "direccion": procedimiento.direccion,
                "tipo_procedimiento": procedimiento.id_tipo_procedimiento.tipo_procedimiento,
                "detalles": descripcion_str,
                "personas_presentes": personas_str,
                "descripcion": descripcion_str,
                "material_utilizado": material_utilizado_str,
                "status": status_str,
                "traslados": traslados_str
            }
        )

    return JsonResponse(datos, safe=False)

def generar_excel_rescate(request):

    mes_excel = request.GET.get('mes') 

    # Extraer año y mes
    año, mes_num = mes_excel.split("-")


    # Optimizar consultas con select_related y prefetch_related
    procedimientos = Procedimientos.objects.filter(id_division=1, fecha__year=año, fecha__month=mes_num).select_related(
        "id_division", "id_solicitante", "id_jefe_comision", "id_municipio", 
        "id_parroquia", "id_tipo_procedimiento"
    ).prefetch_related(
        Prefetch("rescate_set", 
             queryset=Rescate.objects.prefetch_related(
                 Prefetch("rescate_persona_set", to_attr="personas_data"),
                 Prefetch("rescate_animal_set", to_attr="animales_data")
             ), 
             to_attr="rescate_data"),
        Prefetch("abastecimiento_agua_set", to_attr="agua_data"), # Ya
        Prefetch("apoyo_unidades_set", to_attr="apoyo_data"), # Ya
        Prefetch("guardia_prevencion_set", to_attr="guard_data"), # Ys
        Prefetch("falsa_alarma_set", to_attr="falsa_data"), # YA
        Prefetch("atenciones_paramedicas_set", 
             queryset=Atenciones_Paramedicas.objects.prefetch_related(
                 Prefetch("emergencias_medicas_set", 
                          queryset=Emergencias_Medicas.objects.prefetch_related(
                              Prefetch("traslado_set", to_attr="traslados_data")
                          ), 
                          to_attr="emergencias_data"),
                 Prefetch("accidentes_transito_set", 
                          queryset=Accidentes_Transito.objects.prefetch_related(
                              Prefetch("detalles_vehiculos_accidente_set", to_attr="vehiculos_data"),
                              Prefetch("lesionados_set", 
                                       queryset=Lesionados.objects.prefetch_related(
                                           Prefetch("traslado_accidente_set", to_attr="traslados_accidente_data")
                                       ), 
                                       to_attr="lesionados_data")
                          ), 
                          to_attr="accidentes_data")
             ), 
             to_attr="atenciones_data"), # Ya
        Prefetch("servicios_especiales_set", to_attr="especial_data"), # Ya
        Prefetch("incendios_set", 
             queryset=Incendios.objects.prefetch_related(
                 Prefetch("retencion_preventiva_incendios_set", to_attr="retencion_data"),
                 Prefetch("persona_presente_set", to_attr="personas_incendio_data"),
                 Prefetch("detalles_vehiculos_set", to_attr="vehiculos_incendio_data")
             ), 
             to_attr="incendios_data"), # Ya
        Prefetch("mitigacion_riesgos_set", to_attr="mitigacion_data"), # Ya
        Prefetch("evaluacion_riesgo_set", 
             queryset=Evaluacion_Riesgo.objects.prefetch_related(
                 Prefetch("persona_presente_eval_set", to_attr="personas_eval_data")
             ), 
             to_attr="evaluacion_data"),# Ya
        Prefetch("puesto_avanzada_set", to_attr="puesto_data"), # Ya
        Prefetch("artificios_pirotecnicos_set",
        queryset=Artificios_Pirotecnicos.objects.prefetch_related(
            Prefetch(
                "incendios_art_set",
                queryset=Incendios_Art.objects.prefetch_related(
                    Prefetch("persona_presente_art_set", to_attr="personas_presentes_data"),
                    Prefetch("detalles_vehiculos_art_set", to_attr="vehiculos_data")
                ),
                to_attr="incendios_data"
            ),
            Prefetch("lesionados_art_set", to_attr="lesionados_data"),
            Prefetch("fallecidos_art_set", to_attr="fallecidos_data"),
        ),
        to_attr="artificios_data"), # Ya
        Prefetch("comisiones_set", 
             queryset=Comisiones.objects.select_related("comision"), 
             to_attr="comisiones_data"), # Ya
    )

    # Crear una lista para los datos
    datos = []

    # Procesar los procedimientos
    for procedimiento in procedimientos:
        # Solicitante y jefe de comisión
        if procedimiento.id_solicitante.apellidos == "Externo":
            solicitante = procedimiento.solicitante_externo
        else:
            solicitante = f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}"


        jefe_comision = (f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}")

        # Detalles y personas
        personas_presentes = []
        detalles = []
        descripcion = []
        material_utilizado = []
        status = []
        traslados = []
        vehiculos = []
        comisiones_presentes = []
        retencion_preventiva = []

        # Agregar datos de comisiones presentes
        for comision in procedimiento.comisiones_data:
            comisiones_presentes.append(f"({comision.comision.tipo_comision} - {comision.nombre_oficial} {comision.apellido_oficial} ({comision.cedula_oficial}) - Unidad {comision.nro_unidad} - Cuadrante {comision.nro_cuadrante})")

        for rescate in procedimiento.rescate_data:
            material_utilizado.append(rescate.material_utilizado)
            status.append(rescate.status)

            # Agregar personas presentes si existen
            for persona in rescate.personas_data:
                # Agregar detalles del tipo de rescate
                detalles.append(rescate.tipo_rescate.tipo_rescate)
                material_utilizado.append(rescate.material_utilizado)
                status.append(rescate.status)
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula} - {persona.edad} años - {persona.sexo})")
                descripcion.append(persona.descripcion)

            # Agregar animales rescatados si existen
            for animal in rescate.animales_data:
                detalles.append(f"{rescate.tipo_rescate.tipo_rescate}: {animal.especie}")
                descripcion.append(animal.descripcion)
     
        # Otros conjuntos relacionados (agregar más si aplica)
        for agua in procedimiento.agua_data:
            detalles.append(f"{agua.id_tipo_servicio.nombre_institucion} {agua.ltrs_agua} litros - {agua.personas_atendidas} Personas Atendidas")
            personas_presentes.append(f"{agua.nombres} {agua.apellidos} ({agua.cedula})")
            descripcion.append(agua.descripcion)
            material_utilizado.append(agua.material_utilizado)
            status.append(agua.status)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for apoyo in procedimiento.apoyo_data:
            detalles.append(f"{apoyo.id_tipo_apoyo.tipo_apoyo}: {apoyo.unidad_apoyada}")
            descripcion.append(apoyo.descripcion)
            material_utilizado.append(apoyo.material_utilizado)
            status.append(apoyo.status)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for guardia in procedimiento.guard_data:
            detalles.append(guardia.id_motivo_prevencion.motivo)
            descripcion.append(guardia.descripcion)
            material_utilizado.append(guardia.material_utilizado)
            status.append(guardia.status)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for alarma in procedimiento.falsa_data:
            detalles.append(alarma.motivo_alarma.motivo)
            descripcion.append(alarma.descripcion)
            material_utilizado.append(alarma.material_utilizado)
            status.append(alarma.status)
        
        for atencion in procedimiento.atenciones_data:
            # Agregar detalles del tipo de atención paramédica

            # Agregar emergencias médicas si existen
            for emergencia in atencion.emergencias_data:
                detalles.append(atencion.tipo_atencion)
                personas_presentes.append(f" {emergencia.nombres} {emergencia.apellidos} ({emergencia.cedula}) {emergencia.edad} años - {emergencia.sexo} [{emergencia.idx}]")
                descripcion.append(emergencia.descripcion)
                material_utilizado.append(emergencia.material_utilizado)
                status.append(emergencia.status)

                # Agregar traslados si existen
                for traslado in emergencia.traslados_data:
                    traslados.append(f"Traslado: {traslado.hospital_trasladado} - {traslado.medico_receptor} - {traslado.mpps_cmt}")

              # Agregar datos de accidentes de tránsito
            for accidente in atencion.accidentes_data:
                detalles.append(f"{atencion.tipo_atencion}: {accidente.tipo_de_accidente.tipo_accidente} - {accidente.cantidad_lesionados} Lesionados")
                material_utilizado.append(accidente.material_utilizado)
                status.append(accidente.status)

                # Agregar detalles de vehículos involucrados
                for vehiculo in accidente.vehiculos_data:
                    vehiculos.append(f"({vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas})")

                # Agregar datos de lesionados
                for lesionado in accidente.lesionados_data:
                    personas_presentes.append(f"({lesionado.nombres} {lesionado.apellidos} {lesionado.cedula} - {lesionado.edad} años{lesionado.sexo} [{lesionado.idx}])")
                    descripcion.append(f"({lesionado.descripcion})")

                    # Agregar traslados asociados a lesionados
                    for traslado_acc in lesionado.traslados_accidente_data:
                        traslados.append(f"Traslado: {traslado_acc.hospital_trasladado} - {traslado_acc.medico_receptor} - {traslado_acc.mpps_cmt}")
 
        # Otros conjuntos relacionados (agregar más si aplica)
        for especial in procedimiento.especial_data:
            detalles.append(especial.tipo_servicio.serv_especiales)
            descripcion.append(especial.descripcion)
            material_utilizado.append(especial.material_utilizado)
            status.append(especial.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for incendio in procedimiento.incendios_data:
            # Detalles del incendio
            detalles.append(incendio.id_tipo_incendio.tipo_incendio)
            descripcion.append(incendio.descripcion)
            material_utilizado.append(incendio.material_utilizado)
            status.append(incendio.status)

            # Agregar Retenciones Preventivas (GLP)
            for retencion in incendio.retencion_data:

                retencion_preventiva.append(f"{retencion.tipo_cilindro} - {retencion.capacidad} - {retencion.serial} - {retencion.nro_constancia_retencion} - Propietario: {retencion.nombre} {retencion.apellidos} ({retencion.cedula})")

            # Agregar Personas Presentes en el Incendio
            for persona in incendio.personas_incendio_data:
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula} - {persona.edad} años)")

            # Agregar Detalles de Vehículos Relacionados
            for vehiculo in incendio.vehiculos_incendio_data:
                vehiculos.append(f"{vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas}")

        # Otros conjuntos relacionados (agregar más si aplica)
        for mitigacion in procedimiento.mitigacion_data:
            detalles.append(mitigacion.id_tipo_servicio.tipo_servicio)
            descripcion.append(mitigacion.descripcion)
            material_utilizado.append(mitigacion.material_utilizado)
            status.append(mitigacion.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for evaluacion in procedimiento.evaluacion_data:
            detalles.append(f"{evaluacion.id_tipo_riesgo.tipo_riesgo}: {evaluacion.tipo_estructura}")
            descripcion.append(evaluacion.descripcion)
            material_utilizado.append(evaluacion.material_utilizado)
            status.append(evaluacion.status)

            # Personas presentes en la evaluación
            for persona in evaluacion.personas_eval_data:
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula}) - {persona.telefono}")

        # Otros conjuntos relacionados (agregar más si aplica)
        for puesto in procedimiento.puesto_data:
            detalles.append(puesto.id_tipo_servicio.tipo_servicio)
            descripcion.append(puesto.descripcion)
            material_utilizado.append(puesto.material_utilizado)
            status.append(puesto.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for artificio in procedimiento.artificios_data:
            # Información básica del artificio
            detalles.append(f"{artificio.tipo_procedimiento.tipo} - {artificio.nombre_comercio} - {artificio.rif_comerciante}")

            for incendio in artificio.incendios_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}: {incendio.id_tipo_incendio.tipo_incendio}")
                descripcion.append(incendio.descripcion)
                material_utilizado.append(incendio.material_utilizado)
                status.append(incendio.status)

                for persona in incendio.personas_presentes_data:
                    personas_presentes.append(f"{persona.nombres} {persona.apellidos} ({persona.cedula}) - {persona.edad} años")
                    
                for vehiculo in incendio.vehiculos_data:
                    vehiculos.append(f"{vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas}")


            for lesionado in artificio.lesionados_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}")
                personas_presentes.append(f"{lesionado.nombres} {lesionado.apellidos} ({lesionado.cedula}) - {lesionado.edad} años - {lesionado.sexo}")
                descripcion.append(lesionado.descripcion)
                status.append(lesionado.status)

            for fallecido in artificio.fallecidos_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}: {fallecido.motivo_fallecimiento}")
                personas_presentes.append(f"{fallecido.nombres} {fallecido.apellidos} ({fallecido.cedula}) - {fallecido.edad} años - {fallecido.sexo}")
                descripcion.append(fallecido.descripcion)
                status.append(fallecido.status)
                material_utilizado.append(fallecido.material_utilizado)

        # Combinar datos
        detalles_str = " -- ".join(detalles) or " "
        personas_str = " -- ".join(personas_presentes) or " "
        descripcion_str = " -- ".join(descripcion) or " "
        material_utilizado_str = " -- ".join(material_utilizado) or " "
        status_str = " -- ".join(status) or " "
        traslados_str = " -- ".join(traslados) or " "
        vehiculos_str = " -- ".join(vehiculos) or " "
        comisiones_presentes_str = " -- ".join(comisiones_presentes) or " "
        retencion_preventiva_str = " -- ".join(retencion_preventiva) or " "

        # Agregar datos al JSON
        datos.append({
            "division": procedimiento.id_division.division or " ",
            "solicitante": solicitante,
            "jefe_comision": jefe_comision,
            "municipio": procedimiento.id_municipio.municipio or " ",
            "parroquia": procedimiento.id_parroquia.parroquia or " ",
            "fecha": procedimiento.fecha or " ",
            "hora": procedimiento.hora or " ",
            "direccion": procedimiento.direccion or " ",
            "tipo_procedimiento": procedimiento.id_tipo_procedimiento.tipo_procedimiento or " ",
            "detalles": detalles_str,
            "personas_presentes": personas_str,
            "descripcion": descripcion_str,
            "material_utilizado": material_utilizado_str,
            "status": status_str,
            "traslados": traslados_str,
            "vehiculos": vehiculos_str,
            "comisiones": comisiones_presentes_str,
            "retencion_preventiva": retencion_preventiva_str,
        })

    # Devolver los datos como JSON
    return JsonResponse(datos, safe=False)

def generar_excel_prevencion(request):
    division = 3
    
    mes_excel = request.GET.get('mes') 
    # Extraer año y mes
    año, mes_num = mes_excel.split("-")
    

    # Obtén los procedimientos filtrados por división, utilizando select_related para optimizar las relaciones de uno a uno y prefetch_related para las relaciones de uno a muchos
    procedimientos = Procedimientos.objects.filter(id_division=division, fecha__year=año, fecha__month=mes_num).select_related(
        "id_solicitante", "id_jefe_comision", "id_municipio", "id_parroquia", "id_tipo_procedimiento"
    ).prefetch_related(
        Prefetch("evaluacion_riesgo_set", to_attr="evaluacion_data"), # Ya
        Prefetch("asesoramiento_set", to_attr="asesoramiento_data"),
        Prefetch("inspeccion_prevencion_asesorias_tecnicas_set"),
        Prefetch("inspeccion_habitabilidad_set"),
        Prefetch("inspeccion_otros_set"),
        Prefetch("inspeccion_arbol_set"),
        Prefetch("investigacion_set__investigacion_vehiculo_set"),
        Prefetch("investigacion_set__investigacion_comercio_set"),
        Prefetch("investigacion_set__investigacion_estructura_vivienda_set"),
        Prefetch("reinspeccion_prevencion_set"),
        Prefetch("retencion_preventiva_set"),
        Prefetch("inspeccion_establecimiento_art_set"),
    )

    datos = []
    for procedimiento in procedimientos:
        
        # Obtener solicitante y jefe de comisión

        if procedimiento.id_solicitante.apellidos == "Externo":
            solicitante = procedimiento.solicitante_externo
        else:
            solicitante = f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}"


        jefe_comision = (f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}")

        # Detalles y personas
        personas_presentes = []
        detalles = []
        descripcion = []
        material_utilizado = []
        status = []
        vehiculos = []
        comisiones_presentes = []
        retencion_preventiva = []

        # Otros conjuntos relacionados (agregar más si aplica)
        for evaluacion in procedimiento.evaluacion_data:
            detalles.append(f"{evaluacion.id_tipo_riesgo.tipo_riesgo}: {evaluacion.tipo_estructura}")
            descripcion.append(evaluacion.descripcion)
            material_utilizado.append(evaluacion.material_utilizado)
            status.append(evaluacion.status)

            # Personas presentes en la evaluación
            for persona in evaluacion.personas_eval_data:
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula}) - {persona.telefono}")


         # Procesar los detalles de asesoramiento
        for asesoria in procedimiento.asesoramiento_data:
            detalles.append(f"{asesoria.nombre_comercio} {asesoria.rif_comercio}")
            personas_presentes.append(f"{asesoria.nombres} {asesoria.apellidos} ({asesoria.cedula}) - {asesoria.sexo} [{asesoria.telefono}]")
            descripcion.append(asesoria.descripcion)
            material_utilizado.append(asesoria.material_utilizado)
            status.append(asesoria.status)

        # Procesar Inspección Prevención Asesorías Técnicas
        for asesoria in procedimiento.inspeccion_prevencion_asesorias_tecnicas_set.all():
            detalles.append(f"{asesoria.tipo_inspeccion} - {asesoria.nombre_comercio} ({asesoria.propietario}, {asesoria.cedula_propietario})")
            personas_presentes.append(f"{asesoria.persona_sitio_nombre} {asesoria.persona_sitio_apellido} ({asesoria.persona_sitio_cedula}) [{asesoria.persona_sitio_telefono}]")
            descripcion.append(asesoria.descripcion)
            material_utilizado.append(asesoria.material_utilizado)
            status.append(asesoria.status)

        # Procesar Inspección Habitabilidad
        for habitabilidad in procedimiento.inspeccion_habitabilidad_set.all():
            detalles.append(habitabilidad.tipo_inspeccion)
            personas_presentes.append(f"{habitabilidad.persona_sitio_nombre} {habitabilidad.persona_sitio_apellido} ({habitabilidad.persona_sitio_cedula}) [{habitabilidad.persona_sitio_telefono}]")
            descripcion.append(habitabilidad.descripcion)
            material_utilizado.append(habitabilidad.material_utilizado)
            status.append(habitabilidad.status)

        # Procesar Inspección Otros
        for otros in procedimiento.inspeccion_otros_set.all():
            detalles.append(f"{otros.tipo_inspeccion} - {otros.especifique}")
            personas_presentes.append(f"{otros.persona_sitio_nombre} {otros.persona_sitio_apellido} ({otros.persona_sitio_cedula}) [{otros.persona_sitio_telefono}]")
            descripcion.append(otros.descripcion)
            material_utilizado.append(otros.material_utilizado)
            status.append(otros.status)

        # Procesar Inspección Árbol
        for arbol in procedimiento.inspeccion_arbol_set.all():
            detalles.append(f"{arbol.tipo_inspeccion} - Especie: {arbol.especie}, Altura: {arbol.altura_aprox}, Ubicación: {arbol.ubicacion_arbol}")
            personas_presentes.append(f"{arbol.persona_sitio_nombre} {arbol.persona_sitio_apellido} ({arbol.persona_sitio_cedula}) [{arbol.persona_sitio_telefono}]")
            descripcion.append(arbol.descripcion)
            material_utilizado.append(arbol.material_utilizado)
            status.append(arbol.status)

         # Procesar investigaciones
        
        for investigacion in procedimiento.investigacion_set.all():
            # Procesar investigación vehículo
            for vehiculo in investigacion.investigacion_vehiculo_set.all():
                detalles.append(f"{investigacion.tipo_siniestro} - {investigacion.id_tipo_investigacion.tipo_investigacion}")
                vehiculos.append(f"Vehículo: {vehiculo.marca} {vehiculo.modelo}, {vehiculo.color}, Año {vehiculo.año}, Placas {vehiculo.placas}")
                personas_presentes.append(f"Propietario: {vehiculo.nombre_propietario} {vehiculo.apellido_propietario} ({vehiculo.cedula_propietario})")
                descripcion.append(vehiculo.descripcion)
                material_utilizado.append(vehiculo.material_utilizado)
                status.append(vehiculo.status)

            # Procesar investigación comercio
            for comercio in investigacion.investigacion_comercio_set.all():
                detalles.append(f"{investigacion.tipo_siniestro} - {investigacion.id_tipo_investigacion.tipo_investigacion}, Comercio: {comercio.nombre_comercio} ({comercio.rif_comercio})")
                personas_presentes.append(f"{comercio.nombre_propietario} {comercio.apellido_propietario} ({comercio.cedula_propietario})")
                descripcion.append(comercio.descripcion)
                material_utilizado.append(comercio.material_utilizado)
                status.append(comercio.status)

            # Procesar investigación estructura/vivienda
            for estructura in investigacion.investigacion_estructura_vivienda_set.all():
                detalles.append(f"{investigacion.tipo_siniestro} - {investigacion.id_tipo_investigacion.tipo_investigacion}, Estructura: {estructura.tipo_estructura}")
                personas_presentes.append(f"{estructura.nombre} {estructura.apellido} ({estructura.cedula})")
                descripcion.append(estructura.descripcion)
                material_utilizado.append(estructura.material_utilizado)
                status.append(estructura.status)

        # Procesar Reinspección de Prevención
        for reinspeccion in procedimiento.reinspeccion_prevencion_set.all():
            detalles.append(f"Comercio: {reinspeccion.nombre_comercio} ({reinspeccion.rif_comercio})")
            personas_presentes.append(f"{reinspeccion.nombre} {reinspeccion.apellidos} ({reinspeccion.cedula}) - {reinspeccion.sexo} [{reinspeccion.telefono}]")
            descripcion.append(reinspeccion.descripcion)
            material_utilizado.append(reinspeccion.material_utilizado)
            status.append(reinspeccion.status)

        # Procesar Retención Preventiva (GLP)
        for retencion in procedimiento.retencion_preventiva_set.all():
            detalles.append(f"Cilindro: {retencion.tipo_cilindro}, Capacidad: {retencion.capacidad}, Serial: {retencion.serial}, Constancia: {retencion.nro_constancia_retencion}")
            personas_presentes.append(f"{retencion.nombre} {retencion.apellidos} ({retencion.cedula})")
            descripcion.append(retencion.descripcion)
            material_utilizado.append(retencion.material_utilizado)
            status.append(retencion.status)

        # Procesar Inspección de Establecimiento de Artificios
        for inspeccion in procedimiento.inspeccion_establecimiento_art_set.all():
            detalles.append(f"Comercio: {inspeccion.nombre_comercio} (RIF: {inspeccion.rif_comercio})")
            personas_presentes.append(f"{inspeccion.encargado_nombre} {inspeccion.encargado_apellidos} ({inspeccion.encargado_cedula}) - {inspeccion.encargado_sexo}")
            descripcion.append(inspeccion.descripcion)
            material_utilizado.append(inspeccion.material_utilizado)
            status.append(inspeccion.status)

        # Combinar datos
        detalles_str = " -- ".join(detalles) or " "
        personas_str = " -- ".join(personas_presentes) or " "
        descripcion_str = " -- ".join(descripcion) or " "
        material_utilizado_str = " -- ".join(material_utilizado) or " "
        status_str = " -- ".join(status) or " "
        vehiculos_str = " -- ".join(vehiculos) or " "
        comisiones_presentes_str = " -- ".join(comisiones_presentes) or " "
        retencion_preventiva_str = " -- ".join(retencion_preventiva) or " "

        datos.append({
            "division": procedimiento.id_division.division,
            "solicitante": solicitante,
            "jefe_comision": jefe_comision,
            "municipio": procedimiento.id_municipio.municipio,
            "parroquia": procedimiento.id_parroquia.parroquia,
            "fecha": procedimiento.fecha,
            "hora": procedimiento.hora,
            "direccion": procedimiento.direccion,
            "tipo_procedimiento": procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            "detalles": detalles_str,
            "personas_presentes": personas_str,
            "descripcion": descripcion_str,
            "material_utilizado": material_utilizado_str,
            "status": status_str,
            "vehiculos": vehiculos_str,
            "comisiones": comisiones_presentes_str,
            "retencion_preventiva": retencion_preventiva_str,
        })

    return JsonResponse(datos, safe=False, encoder=DjangoJSONEncoder)

def generar_excel_prehospitalaria(request):
    # Crear una lista para almacenar los datos
    datos = []

    division = 5

    mes_excel = request.GET.get('mes') 

    # Extraer año y mes
    año, mes_num = mes_excel.split("-")

    # Consulta optimizada usando select_related y prefetch_related
    procedimientos = Procedimientos.objects.filter(id_division=division, fecha__year=año, fecha__month=mes_num).select_related(
        'id_division', 'id_solicitante', 'id_jefe_comision', 'id_municipio', 'id_parroquia', 'id_tipo_procedimiento'
    ).prefetch_related(
        Prefetch("apoyo_unidades_set", to_attr="apoyo_data"), # Ya
        Prefetch("guardia_prevencion_set", to_attr="guard_data"), # Ys
        Prefetch("falsa_alarma_set", to_attr="falsa_data"), # YA
        Prefetch("atenciones_paramedicas_set", 
             queryset=Atenciones_Paramedicas.objects.prefetch_related(
                 Prefetch("emergencias_medicas_set", 
                          queryset=Emergencias_Medicas.objects.prefetch_related(
                              Prefetch("traslado_set", to_attr="traslados_data")
                          ), 
                          to_attr="emergencias_data"),
                 Prefetch("accidentes_transito_set", 
                          queryset=Accidentes_Transito.objects.prefetch_related(
                              Prefetch("detalles_vehiculos_accidente_set", to_attr="vehiculos_data"),
                              Prefetch("lesionados_set", 
                                       queryset=Lesionados.objects.prefetch_related(
                                           Prefetch("traslado_accidente_set", to_attr="traslados_accidente_data")
                                       ), 
                                       to_attr="lesionados_data")
                          ), 
                          to_attr="accidentes_data")
             ), 
             to_attr="atenciones_data"), # Ya
        Prefetch("servicios_especiales_set", to_attr="especial_data"), # YO
        Prefetch("atendido_no_efectuado_set", to_attr="atendido_data"), # Ya
        Prefetch("puesto_avanzada_set", to_attr="puesto_data"), # Ya
        Prefetch("comisiones_set", 
             queryset=Comisiones.objects.select_related("comision"), 
             to_attr="comisiones_data"), # Ya
        Prefetch("fallecidos_set", to_attr="fallecido_data"), # Ya
        Prefetch("traslado_prehospitalaria_set", to_attr="trasladado_data"),
    )

    for procedimiento in procedimientos:
        # Solicitante y jefe de comisión
        if procedimiento.id_solicitante.apellidos == "Externo":
            solicitante = procedimiento.solicitante_externo
        else:
            solicitante = f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}"


        jefe_comision = (f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}")

        # Detalles y personas
        personas_presentes = []
        detalles = []
        descripcion = []
        material_utilizado = []
        status = []
        traslados = []
        vehiculos = []
        comisiones_presentes = []
        retencion_preventiva = []

        # Agregar datos de comisiones presentes
        for comision in procedimiento.comisiones_data:
            comisiones_presentes.append(f"({comision.comision.tipo_comision} - {comision.nombre_oficial} {comision.apellido_oficial} ({comision.cedula_oficial}) - Unidad {comision.nro_unidad} - Cuadrante {comision.nro_cuadrante})")

         # Otros conjuntos relacionados (agregar más si aplica)
        for apoyo in procedimiento.apoyo_data:
            detalles.append(f"{apoyo.id_tipo_apoyo.tipo_apoyo}: {apoyo.unidad_apoyada}")
            descripcion.append(apoyo.descripcion)
            material_utilizado.append(apoyo.material_utilizado)
            status.append(apoyo.status)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for guardia in procedimiento.guard_data:
            detalles.append(guardia.id_motivo_prevencion.motivo)
            descripcion.append(guardia.descripcion)
            material_utilizado.append(guardia.material_utilizado)
            status.append(guardia.status)
        
         # Otros conjuntos relacionados (agregar más si aplica)
        for atendido in procedimiento.atendido_data:
            descripcion.append(atendido.descripcion)
            material_utilizado.append(atendido.material_utilizado)
            status.append(atendido.status)

            # Otros conjuntos relacionados (agregar más si aplica)
        for alarma in procedimiento.falsa_data:
            detalles.append(alarma.motivo_alarma.motivo)
            descripcion.append(alarma.descripcion)
            material_utilizado.append(alarma.material_utilizado)
            status.append(alarma.status)

        for atencion in procedimiento.atenciones_data:
            # Agregar detalles del tipo de atención paramédica

            # Agregar emergencias médicas si existen
            for emergencia in atencion.emergencias_data:
                detalles.append(atencion.tipo_atencion)
                personas_presentes.append(f" {emergencia.nombres} {emergencia.apellidos} ({emergencia.cedula}) {emergencia.edad} años - {emergencia.sexo} [{emergencia.idx}]")
                descripcion.append(emergencia.descripcion)
                material_utilizado.append(emergencia.material_utilizado)
                status.append(emergencia.status)

                # Agregar traslados si existen
                for traslado in emergencia.traslados_data:
                    traslados.append(f"Traslado: {traslado.hospital_trasladado} - {traslado.medico_receptor} - {traslado.mpps_cmt}")

              # Agregar datos de accidentes de tránsito
            for accidente in atencion.accidentes_data:
                detalles.append(f"{atencion.tipo_atencion}: {accidente.tipo_de_accidente.tipo_accidente} - {accidente.cantidad_lesionados} Lesionados")
                material_utilizado.append(accidente.material_utilizado)
                status.append(accidente.status)

                # Agregar detalles de vehículos involucrados
                for vehiculo in accidente.vehiculos_data:
                    vehiculos.append(f"({vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas})")

                # Agregar datos de lesionados
                for lesionado in accidente.lesionados_data:
                    personas_presentes.append(f"({lesionado.nombres} {lesionado.apellidos} {lesionado.cedula} - {lesionado.edad} años{lesionado.sexo} [{lesionado.idx}])")
                    descripcion.append(f"({lesionado.descripcion})")

                    # Agregar traslados asociados a lesionados
                    for traslado_acc in lesionado.traslados_accidente_data:
                        traslados.append(f"Traslado: {traslado_acc.hospital_trasladado} - {traslado_acc.medico_receptor} - {traslado_acc.mpps_cmt}")
 
        # Otros conjuntos relacionados (agregar más si aplica)
        for especial in procedimiento.especial_data:
            detalles.append(especial.tipo_servicio.serv_especiales)
            descripcion.append(especial.descripcion)
            material_utilizado.append(especial.material_utilizado)
            status.append(especial.status)

        # Agregar Fallecidos
        for fallecido in procedimiento.fallecido_data:
            detalles.append(fallecido.motivo_fallecimiento)
            personas_presentes.append(f"{fallecido.nombres} {fallecido.apellidos} ({fallecido.cedula}) - {fallecido.edad} años - {fallecido.sexo}") 
            descripcion.append(fallecido.descripcion)
            material_utilizado.append(fallecido.material_utilizado)
            status.append(fallecido.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for puesto in procedimiento.puesto_data:
            detalles.append(puesto.id_tipo_servicio.tipo_servicio)
            descripcion.append(puesto.descripcion)
            material_utilizado.append(puesto.material_utilizado)
            status.append(puesto.status)

        for trasladado in procedimiento.trasladado_data:
            detalles.append(trasladado.id_tipo_traslado.tipo_traslado)
            personas_presentes.append(f"{trasladado.nombre} {trasladado.apellido} ({trasladado.cedula}) - {trasladado.edad} años - {trasladado.sexo} [{trasladado.idx}]")
            traslados.append(f"{trasladado.hospital_trasladado} - {trasladado.medico_receptor} - {trasladado.mpps_cmt}")
            descripcion.append(trasladado.descripcion)
            material_utilizado.append(trasladado.material_utilizado)
            status.append(trasladado.status)

        # Combinar datos
        detalles_str = " -- ".join(detalles) or " "
        personas_str = " -- ".join(personas_presentes) or " "
        descripcion_str = " -- ".join(descripcion) or " "
        material_utilizado_str = " -- ".join(material_utilizado) or " "
        status_str = " -- ".join(status) or " "
        traslados_str = " -- ".join(traslados) or " "
        vehiculos_str = " -- ".join(vehiculos) or " "
        comisiones_presentes_str = " -- ".join(comisiones_presentes) or " "
        retencion_preventiva_str = " -- ".join(retencion_preventiva) or " "

        # Agregar los datos al diccionario
        datos.append({
            "division": procedimiento.id_division.division,
            "solicitante": solicitante,
            "jefe_comision": jefe_comision,
            "municipio": procedimiento.id_municipio.municipio,
            "parroquia": procedimiento.id_parroquia.parroquia,
            "fecha": procedimiento.fecha,
            "hora": procedimiento.hora,
            "direccion": procedimiento.direccion,
            "tipo_procedimiento": procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            "detalles": detalles_str,
            "personas_presentes": personas_str,
            "descripcion": descripcion_str,
            "material_utilizado": material_utilizado_str,
            "status": status_str,
            "traslados": traslados_str,
            "vehiculos": vehiculos_str,
            "comisiones": comisiones_presentes_str,
            "retencion_preventiva": retencion_preventiva_str,
        })

    # Configurar la respuesta HTTP para enviar el archivo JSON
    return JsonResponse(datos, safe=False)

def generar_excel_serviciosmedicos(request):

    division_id = 7  # División correspondiente a "Servicios Médicos"

    mes_excel = request.GET.get('mes') 

    # Extraer año y mes
    año, mes_num = mes_excel.split("-")

    # Optimización de consultas usando `select_related` y `prefetch_related`
    procedimientos = Procedimientos.objects.filter(id_division=division_id, fecha__year=año, fecha__month=mes_num).select_related(
        "id_division",  "id_municipio", 
        "id_parroquia", "id_tipo_procedimiento"
    ).prefetch_related(
        Prefetch("valoracion_medica_set", to_attr="valoracion_data"),
        Prefetch("jornada_medica_set", to_attr="jornada_data"), 
    )

    # Crear una lista para almacenar los datos
    datos = []

    for procedimiento in procedimientos:
        
        personas_presentes = []
        descripcion = []
        material_utilizado = []
        status = []

         # Agregar datos de comisiones presentes
        for valoracion in procedimiento.valoracion_data:
           personas_presentes.append(f"{valoracion.nombre} {valoracion.apellido} [{valoracion.cedula}] {valoracion.edad} años, {valoracion.sexo} ({valoracion.telefono})")
           descripcion.append(valoracion.descripcion)
           material_utilizado.append(valoracion.material_utilizado)
           status.append(valoracion.status)

        for jornada in procedimiento.jornada_data:
            descripcion.append(f"{jornada.nombre_jornada} Personas Atendidas: {jornada.cant_personas_aten} - {jornada.descripcion}")
            material_utilizado.append(jornada.material_utilizado)
            status.append(jornada.status)

        # Agregar datos a la lista
        datos.append({
            "division": procedimiento.id_division.division,
            "tipo_servicio": procedimiento.tipo_servicio,
            "jefe_area": procedimiento.solicitante_externo,
            "municipio": procedimiento.id_municipio.municipio,
            "parroquia": procedimiento.id_parroquia.parroquia,
            "fecha": procedimiento.fecha,
            "hora": procedimiento.hora,
            "direccion": procedimiento.direccion,
            "tipo_procedimiento": procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            "descripcion": " -- ".join(descripcion),
            "material_utilizado": " -- ".join(material_utilizado),
            "status": " -- ".join(status),
            "personas_presentes": " -- ".join(personas_presentes),
        })

    # Devolver los datos como JSON
    return JsonResponse(datos, safe=False)

def generar_excel_psicologia(request):
    division = 8

    mes_excel = request.GET.get('mes') 
    # Extraer año y mes
    año, mes_num = mes_excel.split("-")
    
    # Utilizar select_related para relaciones de uno a uno y prefetch_related para relaciones de muchos a muchos
    procedimientos = Procedimientos.objects.filter(id_division=division, fecha__year=año, fecha__month=mes_num).select_related(
        'id_division', 'id_municipio', 'id_parroquia', 'id_tipo_procedimiento'
    ).prefetch_related(
        Prefetch("procedimientos_psicologia_set", to_attr="psicologia_data"),
    )

    # Preparar lista de datos a enviar
    datos = []

    for procedimiento in procedimientos:
        
        personas_presentes = []
        descripcion = []
        material_utilizado = []
        status = []

        for psicologia in procedimiento.psicologia_data:
            personas_presentes.append(f"{psicologia.nombre} {psicologia.apellido} ({psicologia.cedula}), {psicologia.edad} años [{psicologia.sexo}]")
            descripcion.append(psicologia.descripcion)
            material_utilizado.append(psicologia.material_utilizado)
            status.append(psicologia.status)

        # Convertir listas a cadenas separadas por ' -- '
        personas_presentes_str = " -- ".join(personas_presentes)
        descripcion_str = " -- ".join(descripcion)
        material_utilizado_str = " -- ".join(material_utilizado)
        status_str = " -- ".join(status)
        
        # Agregar datos al arreglo
        datos.append({
            'division': procedimiento.id_division.division,
            'jefe_area': procedimiento.solicitante_externo,
            'municipio': procedimiento.id_municipio.municipio,
            'parroquia': procedimiento.id_parroquia.parroquia,
            'fecha': procedimiento.fecha,
            'hora': procedimiento.hora,
            'direccion': procedimiento.direccion,
            'tipo_procedimiento': procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            'descripcion': descripcion_str,
            'material_utilizado': material_utilizado_str,
            "status": status_str,
            'personas_presentes': personas_presentes_str,
        })

    return JsonResponse({'data': datos})

def generar_excel_grumae(request):
    division = 4
                                                                 
    mes_excel = request.GET.get('mes') 
    # Extraer año y mes
    año, mes_num = mes_excel.split("-")
    
    # Optimizar la consulta usando select_related y prefetch_related
    procedimientos = Procedimientos.objects.filter(id_division=division, fecha__year=año, fecha__month=mes_num).select_related(
        'id_division', 'id_solicitante', 'id_jefe_comision', 'id_municipio', 'id_parroquia', 'id_tipo_procedimiento'
    ).prefetch_related(
        Prefetch("abastecimiento_agua_set", to_attr="agua_data"), # Ya
        Prefetch("apoyo_unidades_set", to_attr="apoyo_data"), # Ya
        Prefetch("guardia_prevencion_set", to_attr="guard_data"), # Ya
        Prefetch("atendido_no_efectuado_set", to_attr="atendido_data"), # Ya
        Prefetch("falsa_alarma_set", to_attr="falsa_data"), # YA
        Prefetch("atenciones_paramedicas_set", 
             queryset=Atenciones_Paramedicas.objects.prefetch_related(
                 Prefetch("emergencias_medicas_set", 
                          queryset=Emergencias_Medicas.objects.prefetch_related(
                              Prefetch("traslado_set", to_attr="traslados_data")
                          ), 
                          to_attr="emergencias_data"),
                 Prefetch("accidentes_transito_set", 
                          queryset=Accidentes_Transito.objects.prefetch_related(
                              Prefetch("detalles_vehiculos_accidente_set", to_attr="vehiculos_data"),
                              Prefetch("lesionados_set", 
                                       queryset=Lesionados.objects.prefetch_related(
                                           Prefetch("traslado_accidente_set", to_attr="traslados_accidente_data")
                                       ), 
                                       to_attr="lesionados_data")
                          ), 
                          to_attr="accidentes_data")
             ), 
             to_attr="atenciones_data"), # Ya
        Prefetch("servicios_especiales_set", to_attr="especial_data"), # Ya
        Prefetch("rescate_set", 
             queryset=Rescate.objects.prefetch_related(
                 Prefetch("rescate_persona_set", to_attr="personas_data"),
                 Prefetch("rescate_animal_set", to_attr="animales_data")
             ), 
             to_attr="rescate_data"),
        Prefetch("incendios_set", 
             queryset=Incendios.objects.prefetch_related(
                 Prefetch("retencion_preventiva_incendios_set", to_attr="retencion_data"),
                 Prefetch("persona_presente_set", to_attr="personas_incendio_data"),
                 Prefetch("detalles_vehiculos_set", to_attr="vehiculos_incendio_data")
             ), 
             to_attr="incendios_data"), # Ya
        Prefetch("fallecidos_set", to_attr="fallecido_data"), # Ya
        Prefetch("mitigacion_riesgos_set", to_attr="mitigacion_data"), # Ya
        Prefetch("evaluacion_riesgo_set", 
             queryset=Evaluacion_Riesgo.objects.prefetch_related(
                 Prefetch("persona_presente_eval_set", to_attr="personas_eval_data")
             ), 
             to_attr="evaluacion_data"),# Ya
        Prefetch("puesto_avanzada_set", to_attr="puesto_data"), # Ya
        Prefetch("artificios_pirotecnicos_set",
        queryset=Artificios_Pirotecnicos.objects.prefetch_related(
            Prefetch(
                "incendios_art_set",
                queryset=Incendios_Art.objects.prefetch_related(
                    Prefetch("persona_presente_art_set", to_attr="personas_presentes_data"),
                    Prefetch("detalles_vehiculos_art_set", to_attr="vehiculos_data")
                ),
                to_attr="incendios_data"
            ),
            Prefetch("lesionados_art_set", to_attr="lesionados_data"),
            Prefetch("fallecidos_art_set", to_attr="fallecidos_data"),
        ),
        to_attr="artificios_data"), # Ya
        Prefetch("comisiones_set", 
             queryset=Comisiones.objects.select_related("comision"), 
             to_attr="comisiones_data"), # Ya
    )

    # Preparar lista de datos a enviar
    datos = []

    for procedimiento in procedimientos:
        # Obtener solicitante y jefe de comisión
        if procedimiento.id_solicitante.apellidos == "Externo":
            solicitante = procedimiento.solicitante_externo
        else:
            solicitante = f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}"


        jefe_comision = (f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}")

        # Detalles y personas
        personas_presentes = []
        detalles = []
        descripcion = []
        material_utilizado = []
        status = []
        traslados = []
        vehiculos = []
        comisiones_presentes = []
        retencion_preventiva = []

        # Agregar datos de comisiones presentes
        for comision in procedimiento.comisiones_data:
            comisiones_presentes.append(f"({comision.comision.tipo_comision} - {comision.nombre_oficial} {comision.apellido_oficial} ({comision.cedula_oficial}) - Unidad {comision.nro_unidad} - Cuadrante {comision.nro_cuadrante})")

         # Otros conjuntos relacionados (agregar más si aplica)
        for agua in procedimiento.agua_data:
            detalles.append(f"{agua.id_tipo_servicio.nombre_institucion} {agua.ltrs_agua} litros - {agua.personas_atendidas} Personas Atendidas")
            personas_presentes.append(f"{agua.nombres} {agua.apellidos} ({agua.cedula})")
            descripcion.append(agua.descripcion)
            material_utilizado.append(agua.material_utilizado)
            status.append(agua.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for apoyo in procedimiento.apoyo_data:
            detalles.append(f"{apoyo.id_tipo_apoyo.tipo_apoyo}: {apoyo.unidad_apoyada}")
            descripcion.append(apoyo.descripcion)
            material_utilizado.append(apoyo.material_utilizado)
            status.append(apoyo.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for guardia in procedimiento.guard_data:
            detalles.append(guardia.id_motivo_prevencion.motivo)
            descripcion.append(guardia.descripcion)
            material_utilizado.append(guardia.material_utilizado)
            status.append(guardia.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for atendido in procedimiento.atendido_data:
            descripcion.append(atendido.descripcion)
            material_utilizado.append(atendido.material_utilizado)
            status.append(atendido.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for alarma in procedimiento.falsa_data:
            detalles.append(alarma.motivo_alarma.motivo)
            descripcion.append(alarma.descripcion)
            material_utilizado.append(alarma.material_utilizado)
            status.append(alarma.status)

        for atencion in procedimiento.atenciones_data:
            # Agregar detalles del tipo de atención paramédica

            # Agregar emergencias médicas si existen
            for emergencia in atencion.emergencias_data:
                detalles.append(atencion.tipo_atencion)
                personas_presentes.append(f" {emergencia.nombres} {emergencia.apellidos} ({emergencia.cedula}) {emergencia.edad} años - {emergencia.sexo} [{emergencia.idx}]")
                descripcion.append(emergencia.descripcion)
                material_utilizado.append(emergencia.material_utilizado)
                status.append(emergencia.status)

                # Agregar traslados si existen
                for traslado in emergencia.traslados_data:
                    traslados.append(f"Traslado: {traslado.hospital_trasladado} - {traslado.medico_receptor} - {traslado.mpps_cmt}")

              # Agregar datos de accidentes de tránsito
            for accidente in atencion.accidentes_data:
                detalles.append(f"{atencion.tipo_atencion}: {accidente.tipo_de_accidente.tipo_accidente} - {accidente.cantidad_lesionados} Lesionados")
                material_utilizado.append(accidente.material_utilizado)
                status.append(accidente.status)

                # Agregar detalles de vehículos involucrados
                for vehiculo in accidente.vehiculos_data:
                    vehiculos.append(f"({vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas})")

                # Agregar datos de lesionados
                for lesionado in accidente.lesionados_data:
                    personas_presentes.append(f"({lesionado.nombres} {lesionado.apellidos} {lesionado.cedula} - {lesionado.edad} años{lesionado.sexo} [{lesionado.idx}])")
                    descripcion.append(f"({lesionado.descripcion})")

                    # Agregar traslados asociados a lesionados
                    for traslado_acc in lesionado.traslados_accidente_data:
                        traslados.append(f"Traslado: {traslado_acc.hospital_trasladado} - {traslado_acc.medico_receptor} - {traslado_acc.mpps_cmt}")
    
         # Otros conjuntos relacionados (agregar más si aplica)
        
        for especial in procedimiento.especial_data:
            detalles.append(especial.tipo_servicio.serv_especiales)
            descripcion.append(especial.descripcion)
            material_utilizado.append(especial.material_utilizado)
            status.append(especial.status)

        for rescate in procedimiento.rescate_data:
            material_utilizado.append(rescate.material_utilizado)
            status.append(rescate.status)

            # Agregar personas presentes si existen
            for persona in rescate.personas_data:
                # Agregar detalles del tipo de rescate
                detalles.append(rescate.tipo_rescate.tipo_rescate)
                material_utilizado.append(rescate.material_utilizado)
                status.append(rescate.status)
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula} - {persona.edad} años - {persona.sexo})")
                descripcion.append(persona.descripcion)

            # Agregar animales rescatados si existen
            for animal in rescate.animales_data:
                detalles.append(f"{rescate.tipo_rescate.tipo_rescate}: {animal.especie}")
                descripcion.append(animal.descripcion)
     
         # Otros conjuntos relacionados (agregar más si aplica)
        
        for incendio in procedimiento.incendios_data:
            # Detalles del incendio
            detalles.append(incendio.id_tipo_incendio.tipo_incendio)
            descripcion.append(incendio.descripcion)
            material_utilizado.append(incendio.material_utilizado)
            status.append(incendio.status)

            # Agregar Retenciones Preventivas (GLP)
            for retencion in incendio.retencion_data:

                retencion_preventiva.append(f"{retencion.tipo_cilindro} - {retencion.capacidad} - {retencion.serial} - {retencion.nro_constancia_retencion} - Propietario: {retencion.nombre} {retencion.apellidos} ({retencion.cedula})")

            # Agregar Personas Presentes en el Incendio
            for persona in incendio.personas_incendio_data:
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula} - {persona.edad} años)")

            # Agregar Detalles de Vehículos Relacionados
            for vehiculo in incendio.vehiculos_incendio_data:
                vehiculos.append(f"{vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas}")

         
         # Agregar Fallecidos
        
        for fallecido in procedimiento.fallecido_data:
            detalles.append(fallecido.motivo_fallecimiento)
            personas_presentes.append(f"{fallecido.nombres} {fallecido.apellidos} ({fallecido.cedula}) - {fallecido.edad} años - {fallecido.sexo}") 
            descripcion.append(fallecido.descripcion)
            material_utilizado.append(fallecido.material_utilizado)
            status.append(fallecido.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for mitigacion in procedimiento.mitigacion_data:
            detalles.append(mitigacion.id_tipo_servicio.tipo_servicio)
            descripcion.append(mitigacion.descripcion)
            material_utilizado.append(mitigacion.material_utilizado)
            status.append(mitigacion.status)

        # Otros conjuntos relacionados (agregar más si aplica)
        for evaluacion in procedimiento.evaluacion_data:
            detalles.append(f"{evaluacion.id_tipo_riesgo.tipo_riesgo}: {evaluacion.tipo_estructura}")
            descripcion.append(evaluacion.descripcion)
            material_utilizado.append(evaluacion.material_utilizado)
            status.append(evaluacion.status)

            # Personas presentes en la evaluación
            for persona in evaluacion.personas_eval_data:
                personas_presentes.append(f"{persona.nombre} {persona.apellidos} ({persona.cedula}) - {persona.telefono}")

        


         # Otros conjuntos relacionados (agregar más si aplica)
        
        # Otros conjuntos relacionados (agregar más si aplica)
        for puesto in procedimiento.puesto_data:
            detalles.append(puesto.id_tipo_servicio.tipo_servicio)
            descripcion.append(puesto.descripcion)
            material_utilizado.append(puesto.material_utilizado)
            status.append(puesto.status)
        
        for artificio in procedimiento.artificios_data:
            # Información básica del artificio
            detalles.append(f"{artificio.tipo_procedimiento.tipo} - {artificio.nombre_comercio} - {artificio.rif_comerciante}")

            for incendio in artificio.incendios_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}: {incendio.id_tipo_incendio.tipo_incendio}")
                descripcion.append(incendio.descripcion)
                material_utilizado.append(incendio.material_utilizado)
                status.append(incendio.status)

                for persona in incendio.personas_presentes_data:
                    personas_presentes.append(f"{persona.nombres} {persona.apellidos} ({persona.cedula}) - {persona.edad} años")
                    
                for vehiculo in incendio.vehiculos_data:
                    vehiculos.append(f"{vehiculo.modelo} {vehiculo.marca} {vehiculo.color} {vehiculo.año} - {vehiculo.placas}")


            for lesionado in artificio.lesionados_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}")
                personas_presentes.append(f"{lesionado.nombres} {lesionado.apellidos} ({lesionado.cedula}) - {lesionado.edad} años - {lesionado.sexo}")
                descripcion.append(lesionado.descripcion)
                status.append(lesionado.status)

            for fallecido in artificio.fallecidos_data:
                detalles.append(f"{artificio.nombre_comercio} - {artificio.rif_comerciante} - {artificio.tipo_procedimiento.tipo}: {fallecido.motivo_fallecimiento}")
                personas_presentes.append(f"{fallecido.nombres} {fallecido.apellidos} ({fallecido.cedula}) - {fallecido.edad} años - {fallecido.sexo}")
                descripcion.append(fallecido.descripcion)
                status.append(fallecido.status)
                material_utilizado.append(fallecido.material_utilizado)

        # Combinar datos
        detalles_str = " -- ".join(detalles) or " "
        personas_str = " -- ".join(personas_presentes) or " "
        descripcion_str = " -- ".join(descripcion) or " "
        material_utilizado_str = " -- ".join(material_utilizado) or " "
        status_str = " -- ".join(status) or " "
        traslados_str = " -- ".join(traslados) or " "
        vehiculos_str = " -- ".join(vehiculos) or " "
        comisiones_presentes_str = " -- ".join(comisiones_presentes) or " "
        retencion_preventiva_str = " -- ".join(retencion_preventiva) or " "

        # Agregar datos al arreglo
        datos.append({
            'division': procedimiento.id_division.division,
            'solicitante': solicitante,
            'jefe_comision': jefe_comision,
            'municipio': procedimiento.id_municipio.municipio,
            'parroquia': procedimiento.id_parroquia.parroquia,
            'fecha': procedimiento.fecha,
            'hora': procedimiento.hora,
            'direccion': procedimiento.direccion,
            'tipo_procedimiento': procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            "detalles": detalles_str,
            "personas_presentes": personas_str,
            "descripcion": descripcion_str,
            "material_utilizado": material_utilizado_str,
            "status": status_str,
            "traslados": traslados_str,
            "vehiculos": vehiculos_str,
            "comisiones": comisiones_presentes_str,
            "retencion_preventiva": retencion_preventiva_str,
        })

    return JsonResponse({'data': datos})

def generar_excel_capacitacion(request):
    division = 9
                                                                                                                              
    mes_excel = request.GET.get('mes') 
    # Extraer año y mes
    año, mes_num = mes_excel.split("-")

    # Obtener los procedimientos de la división especificada con prefetching
    # Obtener datos de los procedimientos con select_related y prefetch_related
    procedimientos = Procedimientos.objects.filter(id_division=division, fecha__year=año, fecha__month=mes_num).select_related(
        'id_solicitante', 'id_jefe_comision', 'id_municipio', 'id_parroquia', 'id_tipo_procedimiento'
    ).prefetch_related(
        Prefetch("procedimientos_capacitacion_set", to_attr="capacitacion_data"),
        Prefetch("procedimientos_brigada_set", to_attr="brigada_data"),
        Prefetch("procedimientos_frente_preventivo_set", to_attr="frente_data"),
    )

    datos = []
    for procedimiento in procedimientos:
        
        # Solicitante y jefe de comisión
        if procedimiento.id_solicitante.apellidos == "Externo":
            solicitante = procedimiento.solicitante_externo
        else:
            solicitante = f"{procedimiento.id_solicitante.jerarquia} {procedimiento.id_solicitante.nombres} {procedimiento.id_solicitante.apellidos}"


        jefe_comision = (f"{procedimiento.id_jefe_comision.jerarquia} {procedimiento.id_jefe_comision.nombres} {procedimiento.id_jefe_comision.apellidos}")

        descripcion = []
        material_utilizado = []
        status = []
        detalles = []

        for capacitacion in procedimiento.capacitacion_data:
            detalles.append(f"{capacitacion.tipo_capacitacion} - Clasificacion: {capacitacion.tipo_clasificacion}, {capacitacion.personas_beneficiadas} Personas Benedificiadas")
            descripcion.append(capacitacion.descripcion)
            material_utilizado.append(capacitacion.material_utilizado)
            status.append(capacitacion.status)
        
        for brigada in procedimiento.brigada_data:
            detalles.append(f"{brigada.tipo_capacitacion} - Clasificacion: {brigada.tipo_clasificacion}, {brigada.personas_beneficiadas} Personas Benedificiadas")
            descripcion.append(brigada.descripcion)
            material_utilizado.append(brigada.material_utilizado)
            status.append(brigada.status)

        for frente in procedimiento.frente_data:
            detalles.append(f"Actividad: {frente.nombre_actividad} - Estrategia: {frente.estrategia}, {frente.personas_beneficiadas} Personas Beneficiadas")
            descripcion.append(frente.descripcion)
            material_utilizado.append(frente.material_utilizado)
            status.append(frente.status)

        # Combinar datos
        detalles_str = " -- ".join(detalles) or " "
        descripcion_str = " -- ".join(descripcion) or " "
        material_utilizado_str = " -- ".join(material_utilizado) or " "
        status_str = " -- ".join(status) or " "

        datos.append({
            "division": procedimiento.id_division.division,
            "solicitante": solicitante,
            "jefe_comision": jefe_comision,
            "municipio": procedimiento.id_municipio.municipio,
            "parroquia": procedimiento.id_parroquia.parroquia,
            "fecha": procedimiento.fecha,
            "hora": procedimiento.hora,
            "direccion": procedimiento.direccion,
            "tipo_procedimiento": procedimiento.id_tipo_procedimiento.tipo_procedimiento,
            "dependencia": procedimiento.dependencia,
            "detalles": detalles_str,
            "descripcion": descripcion_str,
            "material_utilizado": material_utilizado_str,
            "status": status_str,
        })

    return JsonResponse(datos, safe=False)
